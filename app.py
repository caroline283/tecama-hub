import streamlit as st
import pandas as pd
import pdfplumber
import re
import io
import os
import unicodedata
from streamlit_gsheets import GSheetsConnection
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side, Font

# --- 1. CONFIGURAÇÃO DA PÁGINA ---
st.set_page_config(page_title="Tecama Hub Industrial", layout="wide", page_icon="🏗️")

# --- 2. CSS PERSONALIZADO (VISUAL v6.6 - INTEGRAL E TRAVADO) ---
st.markdown("""
    <style>
    [data-testid="stSidebar"] .stRadio div[role="radiogroup"] label { font-size: 22px !important; font-weight: 600 !important; color: #333 !important; }
    h1 { color: #FF5722 !important; font-family: 'Segoe UI', sans-serif; }
    h3 { color: #444 !important; }
    .stButton > button {
        background-color: #FF5722; color: white; width: 100%; border-radius: 12px;
        font-weight: bold; height: 3.5em; font-size: 16px; border: none;
        box-shadow: 0 4px 8px rgba(0,0,0,0.1);
    }
    .home-link .stButton > button {
        background-color: transparent !important; color: #FF5722 !important; border: none !important;
        font-size: 24px !important; font-weight: bold !important; text-align: left !important;
        padding: 0 !important; height: auto !important; text-decoration: underline !important;
    }
    </style>
    """, unsafe_allow_html=True)

conn = st.connection("gsheets", type=GSheetsConnection)

# --- 3. FUNÇÕES AUXILIARES ---
def norm(t):
    if t is None or pd.isna(t): return ""
    t = unicodedata.normalize("NFD", str(t).upper()).encode("ascii", "ignore").decode("utf-8")
    return " ".join(t.split()).strip()

def limpar_material_apenas_cor(t):
    """Filtro solicitado: Deixa apenas o nome da cor"""
    t = norm(t)
    t = re.sub(r'\d+\s*MM', '', t) 
    for r in ["CHAPA DE", "CHAPA", "MDF", "MDP", "HDF", "MM", "DURATEX", "ARACO"]:
        t = t.replace(r, "")
    return t.strip()

def calcular_pesos_madeira(larg, comp, quant, material_texto):
    PESO_M2_BASE = {"MDP": 12.0, "MDF": 13.5}
    try:
        l, c, q = float(str(larg).replace(',','.')), float(str(comp).replace(',','.')), float(str(quant).replace(',','.'))
        m_norm = norm(material_texto)
        tipo = "MDF" if "MDF" in m_norm else "MDP"
        esp = float(re.search(r"(\d+)\s*MM", m_norm).group(1)) if re.search(r"(\d+)\s*MM", m_norm) else 18.0
        p_u = (l/1000) * (c/1000) * PESO_M2_BASE[tipo] * (esp/18)
        return round(p_u, 2), round(p_u * q, 2)
    except: return 0.0, 0.0

# --- 4. NAVEGAÇÃO ---
if 'nav' not in st.session_state: st.session_state.nav = "🏠 Início"

with st.sidebar:
    if os.path.exists("logo_tecama.png"): st.image("logo_tecama.png", use_container_width=True)
    opcao = st.radio("NAVEGAÇÃO", ["🏠 Início", "🌲 Marcenaria", "⚙️ Metalurgia"], 
                     index=["🏠 Início", "🌲 Marcenaria", "⚙️ Metalurgia"].index(st.session_state.nav))
    st.session_state.nav = opcao

# ==========================================
# PÁGINA: INÍCIO (v6.6 TOTALMENTE RESTAURADA)
# ==========================================
if st.session_state.nav == "🏠 Início":
    st.title("Tecama Hub Industrial")
    st.markdown("### Bem-vindo ao Sistema Unificado de Produção")
    st.write("Esta plataforma foi desenvolvida para centralizar as operações das divisões de **Marcenaria** e **Metalurgia**, garantindo agilidade no processamento de pedidos e precisão nos cálculos de engenharia.")
    st.markdown("---")
    st.markdown('<div class="home-link">', unsafe_allow_html=True)
    if st.button("🌲 Divisão de Marcenaria"): st.session_state.nav = "🌲 Marcenaria"; st.rerun()
    st.markdown('</div>', unsafe_allow_html=True)
    st.write("Processamento de arquivos CSV (Pontta) e geração de arquivos para o **Corte Certo**.")
    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown('<div class="home-link">', unsafe_allow_html=True)
    if st.button("⚙️ Divisão de Metalurgia"): st.session_state.nav = "⚙️ Metalurgia"; st.rerun()
    st.markdown('</div>', unsafe_allow_html=True)
    st.write("Levantamento automático de peso através do relatório PDF.")

# ==========================================
# PÁGINA: MARCENARIA
# ==========================================
elif st.session_state.nav == "🌲 Marcenaria":
    st.header("🌲 Marcenaria")
    aba_prod, aba_cc, aba_cores = st.tabs(["📋 Fase 1: Produção", "🚀 Fase 2: Corte Certo", "🎨 Tabelas de Cores"])
    
    with aba_prod:
        up_csv = st.file_uploader("Suba o CSV original do Pontta", type="csv")
        if up_csv:
            df_b = pd.read_csv(up_csv, sep=None, engine='python', dtype=str)
            df_b.columns = [norm(c) for c in df_b.columns]
            if st.button("🚀 Gerar Excel para Fábrica"):
                # Ajuste: Limpa material e calcula pesos
                df_b["MATERIAL"] = df_b["MATERIAL"].apply(limpar_apenas_cor)
                pesos = df_b.apply(lambda r: calcular_pesos_madeira(r.get("LARG",0), r.get("COMP",0), r.get("QUANT",0), r.get("MATERIAL","")), axis=1)
                df_b["PESO_UNIT"] = pesos.apply(lambda x: x[0]); df_b["PESO_TOTAL"] = pesos.apply(lambda x: x[1])
                
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine="openpyxl") as writer:
                    ws = writer.book.create_sheet("PRODUCAO")
                    ws.cell(row=1, column=1, value="TECAMA | PRODUÇÃO").font = Font(bold=True, size=14)
                    header = ["QUANT","COMP","LARG","COR","COD","DESCPECA","PRODUTO","CORTE","FITA","USINAGEM","PESO UNIT.","PESO TOTAL"]
                    for i, h in enumerate(header, 1):
                        cell = ws.cell(row=3, column=i, value=h); cell.font = Font(bold=True); cell.alignment = Alignment(horizontal="center")
                    
                    curr = 4
                    df_b = df_b.sort_values(by="DES_PAI")
                    for prod, g in df_b.groupby("DES_PAI", sort=False):
                        ini = curr
                        for _, r in g.iterrows():
                            vals = [r.get("QUANT"), r.get("COMP"), r.get("LARG"), r.get("MATERIAL"), r.get("COR"), r.get("DESCPECA"), r.get("DES_PAI"), "","","", r.get("PESO_UNIT"), r.get("PESO_TOTAL")]
                            for i, v in enumerate(vals, 1):
                                c = ws.cell(row=curr, column=i, value=v)
                                c.border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
                                # Ajuste: Mesclagem e Quebra de Texto no Produto
                                if i == 7: c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                                else: c.alignment = Alignment(horizontal="center", vertical="center")
                            curr += 1
                        if len(g) > 1: ws.merge_cells(start_row=ini, end_row=curr-1, start_column=7, end_column=7)
                        curr += 1 
                    ws.column_dimensions['G'].width = 30 # Largura média para Produto
                    for i in range(1, 13): 
                        if i != 7: ws.column_dimensions[get_column_letter(i)].width = 15
                st.download_button("📥 Baixar Excel de Produção", output.getvalue(), "PRODUCAO_TECAMA.xlsx")

    with aba_cc:
        up_ex = st.file_uploader("Suba o Excel que você editou manualmente", type="xlsx")
        if up_ex:
            if st.button("🚀 Gerar CSV para Corte Certo"):
                df_e = pd.read_excel(up_ex, skiprows=2).dropna(subset=['QUANT', 'COMP', 'LARG'], how='all')
                res = pd.DataFrame()
                res["ITEM"] = range(1, len(df_e) + 1)
                for c in ["QUANT", "COMP", "LARG"]: 
                    res[c] = pd.to_numeric(df_e[c], errors='coerce').fillna(0).astype(int)
                # Pega a coluna COD (5ª coluna do Excel gerado)
                res["COR"] = df_e["COD"].apply(lambda x: str(int(float(x))) if str(x).replace('.','').isdigit() else str(x))
                res["DESC"] = df_e["DESCPECA"]
                csv_out = res.to_csv(index=False, sep=";", header=False, encoding="utf-8-sig")
                st.download_button("📥 Baixar CSV para Corte Certo", csv_out, "CORTE_CERTO_TECAM.csv")

    with aba_cores:
        df_cores_gs = conn.read(worksheet="CORES_MARCENARIA", ttl=0)
        st.subheader("🎨 Gestão de Códigos de Cores")
        st.data_editor(df_cores_gs, num_rows="dynamic", use_container_width=True, key="editor_cores")

# ==========================================
# PÁGINA: METALURGIA (v6.6 INTEGRAL)
# ==========================================
elif st.session_state.nav == "⚙️ Metalurgia":
    st.header("⚙️ Metalurgia")
    aba_calc_m, aba_db_m = st.tabs(["📋 Calculadora", "🛠️ Tabelas Base"])
    
    try:
        db_map = conn.read(worksheet="MAPEAMENTO_TIPO", ttl=5)
        db_metro = conn.read(worksheet="PESO_POR_METRO", ttl=5)
        db_conj = conn.read(worksheet="PESO_CONJUNTO", ttl=5)
        dict_m = dict(zip(db_metro['secao'].apply(norm), db_metro['peso_kg_m']))
        list_m = db_map.to_dict('records'); list_c = db_conj.to_dict('records')
    except: st.error("Erro ao carregar tabelas base.")

    with aba_calc_m:
        up_pdf = st.file_uploader("Suba o PDF do Relatório", type="pdf")
        if up_pdf:
            itens = []
            with pdfplumber.open(up_pdf) as pdf:
                for page in pdf.pages:
                    for table in page.extract_tables():
                        for r in table:
                            if r and len(r) > 3 and str(r[0]).strip().isdigit():
                                itens.append({"QTD": r[0], "DESCRIÇÃO": r[1], "MEDIDA": r[3], "COR": r[2]})
            df_ed_m = st.data_editor(pd.DataFrame(itens), use_container_width=True)
            if st.button("🚀 Calcular Pesos"):
                res_m = []
                for _, r in df_ed_m.iterrows():
                    desc_l = norm(str(r.get('DESCRIÇÃO')))
                    qtd = float(str(r.get('QTD', 0)).replace(',','.'))
                    tipo = "DESCONHECIDO"
                    for regra in list_m:
                        if norm(regra.get('texto_contido')) in desc_l:
                            tipo = str(regra.get('tipo', 'DESCONHECIDO')).upper(); break
                    if tipo == "IGNORAR": continue
                    p_u = 0.0
                    if tipo == "CONJUNTO":
                        for c in list_c:
                            if norm(c.get('nome_conjunto')) in desc_l: p_u = float(c.get('peso_unit_kg', 0)); break
                    elif "TUBO" in tipo or tipo in dict_m:
                        med = float(str(r.get('MEDIDA', '0')).lower().replace('mm','').replace(',','.').strip())
                        sec = norm(tipo.replace('TUBO ', '').strip())
                        p_u = (med / 1000) * dict_m.get(sec, 0.0)
                    res_m.append({"QTD": qtd, "DESCRIÇÃO": r.get('DESCRIÇÃO'), "MEDIDA": r.get('MEDIDA'), "TIPO": tipo, "PESO UNIT.": round(p_u, 3), "PESO TOTAL": round(p_u * qtd, 3)})
                df_res_m = pd.DataFrame(res_m)
                st.metric("Total Geral", f"{df_res_m['PESO TOTAL'].sum():.2f} kg")
                st.dataframe(df_res_m, use_container_width=True)

    with aba_db_m:
        if 't_ativa_m' not in st.session_state: st.session_state.t_ativa_m = "MAPEAMENTO_TIPO"
        c1, c2, c3 = st.columns(3)
        if c1.button("📋 Mapeamento"): st.session_state.t_ativa_m = "MAPEAMENTO_TIPO"
        if c2.button("⚖️ Tubos"): st.session_state.t_ativa_m = "PESO_POR_METRO"
        if c3.button("📦 Conjuntos"): st.session_state.t_ativa_m = "PESO_CONJUNTO"
        df_view = conn.read(worksheet=st.session_state.t_ativa_m, ttl=0)
        st.subheader(f"Tabela: {st.session_state.t_ativa_m}")
        st.data_editor(df_view, num_rows="dynamic", use_container_width=True)
